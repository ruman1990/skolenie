# letecka_preprava.py

# 1. Vytvorenie letov pomocou namedtuple
from collections import namedtuple

Let = namedtuple("Let", ["cislo", "odlet", "ciel", "pocet_pasazierov"])

lety_namedtuple = [
    Let("SK1001", "Bratislava", "Londýn", 158),
    Let("SK1002", "Košice", "Praha", 110),
    Let("SK1003", "Bratislava", "Paríž", 175),
    Let("SK1004", "Poprad", "Viedeň", 95),
    Let("SK1005", "Bratislava", "Rím", 185),
]
print("1. Namedtuple lety:")
for let in lety_namedtuple:
    print(let)
print("="*40)

# 2. Prevod na dataclass, pridanie typu lietadla
from dataclasses import dataclass

@dataclass
class LetDataclass:
    cislo: str
    odlet: str
    ciel: str
    pocet_pasazierov: int
    typ_lietadla: str

lety_dataclass = [
    LetDataclass("SK1001", "Bratislava", "Londýn", 158, "Boeing 737"),
    LetDataclass("SK1002", "Košice", "Praha", 110, "Airbus A320"),
    LetDataclass("SK1003", "Bratislava", "Paríž", 175, "Boeing 737"),
    LetDataclass("SK1004", "Poprad", "Viedeň", 95, "Embraer 195"),
    LetDataclass("SK1005", "Bratislava", "Rím", 185, "Boeing 737"),
]
print("2. Dataclass lety:")
for let in lety_dataclass:
    print(let)
print("="*40)

# 3. Uloženie do Excelu (openpyxl)
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Lety"
ws.append(["Číslo letu", "Odlet", "Cieľ", "Počet pasažierov", "Typ lietadla"])

for let in lety_dataclass:
    ws.append([let.cislo, let.odlet, let.ciel, let.pocet_pasazierov, let.typ_lietadla])

wb.save("letecka_preprava.xlsx")
print("3. Excel súbor 'letecka_preprava.xlsx' vytvorený.")
print("="*40)

# 4. Načítanie z Excelu, export do CSV (tablib)
import tablib
from openpyxl import load_workbook

wb2 = load_workbook("letecka_preprava.xlsx")
ws2 = wb2["Lety"]

# Prečítaj všetky riadky z Excelu (vynechaj hlavičku)
data = tablib.Dataset()
rows = list(ws2.iter_rows(values_only=True))
data.headers = rows[0]
for row in rows[1:]:
    data.append(row)

with open("letecka_preprava.csv", "w", encoding="utf-8") as f:
    f.write(data.export("csv"))
print("4. CSV súbor 'letecka_preprava.csv' vytvorený.")
print("="*40)

# 5. Sčítanie pasažierov z CSV cez tablib
with open("letecka_preprava.csv", "r", encoding="utf-8") as f:
    data_csv = tablib.import_set(f.read(), format="csv")
celkovy_pocet = sum(int(radek["Počet pasažierov"]) for radek in data_csv.dict)
print(f"5. Celkový počet pasažierov na všetkých letoch je: {celkovy_pocet}")
