from openpyxl import Workbook
import requests
resp = requests.get('https://jsonplaceholder.typicode.com/users')
data = resp.json()

vyfiltrovane = [{"name" : x['name'],"email" : x['email']} for x in data]

wb = Workbook()
ws = wb.active



ws['A1'] = "MENO"
ws['B1'] = "EMAIL"

for x in vyfiltrovane:
    ws.append(list(x.values()))

wb.save('priklad.xlsx')